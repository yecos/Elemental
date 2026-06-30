"""
ELEMENTAL — Calendario de Contenido XLSX
Plan de redes sociales de 4 semanas con paleta dark premium.
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime, timedelta

# ============= BRAND TOKENS =============
LIMA = "C8D400"
LIMA_DARK = "B5C200"
NEGRO = "0D0D0D"
CARBON = "4A4A4A"
GRIS_CLARO = "F2F2F2"
BLANCO = "FFFFFF"
SURFACE = "1A1A1A"
SURFACE_2 = "262626"
BORDER_C = "333333"

# Fills
fill_negro = PatternFill("solid", fgColor=NEGRO)
fill_surface = PatternFill("solid", fgColor=SURFACE)
fill_surface_2 = PatternFill("solid", fgColor=SURFACE_2)
fill_lima = PatternFill("solid", fgColor=LIMA)
fill_gris = PatternFill("solid", fgColor=GRIS_CLARO)
fill_blanco = PatternFill("solid", fgColor=BLANCO)
fill_carbon = PatternFill("solid", fgColor=CARBON)

# Fonts
font_title = Font(name="Space Grotesk", size=22, bold=True, color=BLANCO)
font_subtitle = Font(name="Inter", size=11, color=GRIS_CLARO)
font_header = Font(name="Space Grotesk", size=11, bold=True, color=BLANCO)
font_subheader = Font(name="Space Grotesk", size=10, bold=True, color=LIMA)
font_body = Font(name="Inter", size=10, color=BLANCO)
font_body_dark = Font(name="Inter", size=10, color=NEGRO)
font_muted = Font(name="Inter", size=9, color="8A8A8A")
font_mono = Font(name="JetBrains Mono", size=9, color=LIMA)
font_lima_bold = Font(name="Space Grotesk", size=11, bold=True, color=NEGRO)

# Borders
side_thin = Side(style="thin", color=BORDER_C)
side_lima = Side(style="thin", color=LIMA)
border_all = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
border_left_lima = Border(left=side_lima, right=side_thin, top=side_thin, bottom=side_thin)

# Alignments
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header_row(ws, row, col_start, col_end, fill=fill_negro, font=font_header):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align_center
        cell.border = border_all


def style_body_cell(cell, fill=fill_surface, font=font_body, align=align_top_left):
    cell.fill = fill
    cell.font = font
    cell.alignment = align
    cell.border = border_all


# ============= CREATE WORKBOOK =============
wb = Workbook()
wb.properties.creator = "ELEMENTAL"
wb.properties.title = "Calendario de Contenido - ELEMENTAL"
wb.properties.subject = "Plan de redes sociales 4 semanas"

# Remove default sheet, will add custom ones
wb.remove(wb.active)

# ============= SHEET 1: RESUMEN =============
ws1 = wb.create_sheet("Resumen")
ws1.sheet_view.showGridLines = False
ws1.sheet_view.zoomScale = 110

# Background fill for whole sheet (dark premium)
for r in range(1, 50):
    for c in range(1, 12):
        ws1.cell(row=r, column=c).fill = fill_negro

# Title row
ws1["B2"] = "ELEMENTAL"
ws1["B2"].font = Font(name="Space Grotesk", size=28, bold=True, color=BLANCO)
ws1["B2"].fill = fill_negro

ws1["B3"] = "."
ws1["B3"].font = Font(name="Space Grotesk", size=28, bold=True, color=LIMA)
ws1["B3"].fill = fill_negro

ws1["B4"] = "Calendario de contenido · Redes sociales · 4 semanas"
ws1["B4"].font = Font(name="JetBrains Mono", size=10, color=LIMA)
ws1["B4"].fill = fill_negro

ws1["B5"] = "Plan estratégico de contenido para Instagram, LinkedIn y WhatsApp. Cadencia: 3 carruseles + 3 reels + 7 stories + 1 caso de éxito por semana."
ws1["B5"].font = Font(name="Inter", size=11, color=GRIS_CLARO)
ws1["B5"].fill = fill_negro
ws1["B5"].alignment = align_top_left
ws1.row_dimensions[5].height = 36
ws1.merge_cells("B5:K5")

# Section: Estadísticas
ws1["B7"] = "ESTADÍSTICAS DEL PLAN"
ws1["B7"].font = font_subheader
ws1["B7"].fill = fill_negro

stats = [
    ("Semanas", "4"),
    ("Carruseles", "12"),
    ("Reels", "12"),
    ("Stories", "28"),
    ("Casos de éxito", "4"),
    ("Total publicaciones", "56"),
    ("Pilares de contenido", "5"),
    ("Formatos activos", "5"),
]
for i, (k, v) in enumerate(stats, start=8):
    ws1.cell(row=i, column=2, value=k).font = font_body
    ws1.cell(row=i, column=2).fill = fill_surface
    ws1.cell(row=i, column=2).border = border_all
    ws1.cell(row=i, column=2).alignment = align_left
    ws1.cell(row=i, column=3, value=v).font = font_lima_bold
    ws1.cell(row=i, column=3).fill = fill_surface
    ws1.cell(row=i, column=3).border = border_all
    ws1.cell(row=i, column=3).alignment = align_center

# Section: Pilares
ws1["B17"] = "PILARES DE CONTENIDO"
ws1["B17"].font = font_subheader
ws1["B17"].fill = fill_negro

pillars = [
    ("01", "IA para empresas", "Prompts, bots, automatizaciones, casos prácticos y tendencias aplicadas."),
    ("02", "Marketing y ventas", "Embudos, anuncios, contenido, SEO, redes y estrategias de conversión."),
    ("03", "Antes / después", "Transformaciones de marca, web, piezas, campañas y procesos."),
    ("04", "Casos reales", "Resultados, aprendizajes y mini auditorías de empresas."),
    ("05", "Cultura ELEMENTAL", "Proceso creativo, metodología, equipo, herramientas y backstage."),
]
for i, (num, name, desc) in enumerate(pillars, start=18):
    ws1.cell(row=i, column=2, value=num).font = Font(name="Space Grotesk", size=11, bold=True, color=LIMA)
    ws1.cell(row=i, column=2).fill = fill_surface
    ws1.cell(row=i, column=2).border = border_all
    ws1.cell(row=i, column=2).alignment = align_center
    ws1.cell(row=i, column=3, value=name).font = Font(name="Space Grotesk", size=11, bold=True, color=BLANCO)
    ws1.cell(row=i, column=3).fill = fill_surface
    ws1.cell(row=i, column=3).border = border_all
    ws1.cell(row=i, column=3).alignment = align_left
    ws1.cell(row=i, column=4, value=desc).font = font_body
    ws1.cell(row=i, column=4).fill = fill_surface
    ws1.cell(row=i, column=4).border = border_all
    ws1.cell(row=i, column=4).alignment = align_top_left
    ws1.merge_cells(start_row=i, start_column=4, end_row=i, end_column=11)
    ws1.row_dimensions[i].height = 30

# Section: KPIs objetivo
ws1["B25"] = "KPIs OBJETIVO · 30 DÍAS"
ws1["B25"].font = font_subheader
ws1["B25"].fill = fill_negro

kpis = [
    ("Alcance mensual Instagram", "80,000 - 120,000"),
    ("Impresiones totales", "150,000+"),
    ("Nuevos seguidores", "+400 a +800"),
    ("Tasa de engagement", "3.5% - 5%"),
    ("Guardados por carrusel", "50+"),
    ("Leads por WhatsApp", "30 - 60"),
    ("Diagnósticos agendados", "8 - 15"),
]
for i, (k, v) in enumerate(kpis, start=26):
    ws1.cell(row=i, column=2, value=k).font = font_body
    ws1.cell(row=i, column=2).fill = fill_surface
    ws1.cell(row=i, column=2).border = border_all
    ws1.cell(row=i, column=2).alignment = align_left
    ws1.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
    ws1.cell(row=i, column=7, value=v).font = font_lima_bold
    ws1.cell(row=i, column=7).fill = fill_surface
    ws1.cell(row=i, column=7).border = border_all
    ws1.cell(row=i, column=7).alignment = align_center
    ws1.merge_cells(start_row=i, start_column=7, end_row=i, end_column=11)

# Section: Cadencia
ws1["B35"] = "CADENCIA SEMANAL"
ws1["B35"].font = font_subheader
ws1["B35"].fill = fill_negro

cadence = [
    ("Lunes", "Carrusel educativo", "Pilar 1 o 2"),
    ("Martes", "Reel corto (tip)", "Pilar 1"),
    ("Miércoles", "Carrusel caso / auditoría", "Pilar 4"),
    ("Jueves", "Reel corto (error común)", "Pilar 2"),
    ("Viernes", "Carrusel antes/después", "Pilar 3"),
    ("Sábado", "Reel cultura/backstage", "Pilar 5"),
    ("Domingo", "Story + caso de éxito", "Pilar 4"),
]
ws1.cell(row=36, column=2, value="Día").font = font_header
ws1.cell(row=36, column=2).fill = fill_surface_2
ws1.cell(row=36, column=2).border = border_all
ws1.cell(row=36, column=2).alignment = align_center
ws1.cell(row=36, column=3, value="Formato").font = font_header
ws1.cell(row=36, column=3).fill = fill_surface_2
ws1.cell(row=36, column=3).border = border_all
ws1.cell(row=36, column=3).alignment = align_center
ws1.cell(row=36, column=4, value="Pilar").font = font_header
ws1.cell(row=36, column=4).fill = fill_surface_2
ws1.cell(row=36, column=4).border = border_all
ws1.cell(row=36, column=4).alignment = align_center
ws1.merge_cells(start_row=36, start_column=4, end_row=36, end_column=11)

for i, (day, fmt, pillar) in enumerate(cadence, start=37):
    ws1.cell(row=i, column=2, value=day).font = font_lima_bold
    ws1.cell(row=i, column=2).fill = fill_surface
    ws1.cell(row=i, column=2).border = border_all
    ws1.cell(row=i, column=2).alignment = align_center
    ws1.cell(row=i, column=3, value=fmt).font = font_body
    ws1.cell(row=i, column=3).fill = fill_surface
    ws1.cell(row=i, column=3).border = border_all
    ws1.cell(row=i, column=3).alignment = align_left
    ws1.cell(row=i, column=4, value=pillar).font = font_body
    ws1.cell(row=i, column=4).fill = fill_surface
    ws1.cell(row=i, column=4).border = border_all
    ws1.cell(row=i, column=4).alignment = align_left
    ws1.merge_cells(start_row=i, start_column=4, end_row=i, end_column=11)

# Footer
ws1["B45"] = "ELEMENTAL · Diseño. Tecnología. Inteligencia."
ws1["B45"].font = Font(name="JetBrains Mono", size=9, color="8A8A8A")
ws1["B45"].fill = fill_negro

set_col_widths(ws1, [3, 18, 14, 14, 14, 14, 14, 14, 14, 14, 14])

# ============= SHEET 2: CALENDARIO MAESTRO =============
ws2 = wb.create_sheet("Calendario")
ws2.sheet_view.showGridLines = False
ws2.sheet_view.zoomScale = 110
ws2.freeze_panes = "A4"

# Dark background
for r in range(1, 80):
    for c in range(1, 13):
        ws2.cell(row=r, column=c).fill = fill_negro

# Title
ws2["A1"] = "CALENDARIO MAESTRO · 4 SEMANAS"
ws2["A1"].font = Font(name="Space Grotesk", size=14, bold=True, color=BLANCO)
ws2["A1"].fill = fill_negro
ws2.merge_cells("A1:L1")

ws2["A2"] = "Cada fila es una publicación. Filtrar por Semana, Tipo o Pilar según necesidad."
ws2["A2"].font = Font(name="Inter", size=10, color="8A8A8A")
ws2["A2"].fill = fill_negro
ws2.merge_cells("A2:L2")

# Headers
headers = ["Semana", "Día", "Fecha", "Tipo", "Pilar", "Tema", "Título", "Copy / Descripción", "Hashtags", "CTA", "Formato", "Estado"]
for i, h in enumerate(headers, 1):
    cell = ws2.cell(row=3, column=i, value=h)
    cell.fill = fill_surface_2
    cell.font = font_header
    cell.alignment = align_center
    cell.border = border_all
ws2.row_dimensions[3].height = 28

# Content plan — 4 weeks × 7 posts = 28 posts
# Format: (semana, dia, fecha, tipo, pilar, tema, titulo, copy, hashtags, cta, formato, estado)
posts = [
    # ====== SEMANA 1 ======
    (1, "Lun", "2026-01-05", "Carrusel", "01 · IA", "Prompts IA empresas",
     "5 prompts de IA para vender más en WhatsApp",
     "Carrusel 6 slides con 5 prompts listos para usar en WhatsApp Business: clasificar leads, redactar respuestas, generar cotizaciones, resumir conversaciones y detectar oportunidades. Cada slide: prompt + ejemplo + tip.",
     "#IA #WhatsAppBusiness #PromptsIA #Automatizacion #Ventas",
     "Guarda este post",
     "Carrusel 1080×1350", "Programado"),
    (1, "Mar", "2026-01-06", "Reel", "01 · IA", "Bot WhatsApp en 60s",
     "Cómo montar un bot de WhatsApp en 60 segundos",
     "Reel vertical 30s. Demo en pantalla dividida: izquierda código/no-code, derecha resultado. Audio trending + voz explicando 3 pasos clave. Hook en primeros 3s.",
     "#IA #WhatsAppBot #Automatizacion #Tutorial",
     "Comenta 'BOT' y te envío la guía",
     "Reel 1080×1920", "Programado"),
    (1, "Mié", "2026-01-07", "Carrusel", "04 · Casos", "Mini auditoría restaurante",
     "Mini auditoría: restaurante que perdió 180 reservas en 2 meses",
     "Carrusel 7 slides. Slide 1 hook con número impactante. Slides 2-5: 4 errores detectados (sin WhatsApp, web lenta, sin seguimiento, sin reservas online). Slide 6: solución propuesta. Slide 7: CTA a diagnóstico gratuito.",
     "#Auditoria #Restaurantes #MarketingDigital #Casos",
     "Agenda tu auditoría",
     "Carrusel 1080×1350", "Programado"),
    (1, "Jue", "2026-01-08", "Reel", "02 · Marketing", "Error en Meta Ads",
     "El error que te hace perder dinero en Meta Ads",
     "Reel 25s. Persona mirando cámara, fondo oscuro. Explica el error #1: audiences frías sin warm-up. Muestra 1 fix concreto. Hook: 'Si estás haciendo esto, pierdes plata.'",
     "#MetaAds #Marketing #PublicidadDigital",
     "Comenta ADS para guía completa",
     "Reel 1080×1920", "Programado"),
    (1, "Vie", "2026-01-09", "Carrusel", "03 · Antes/Después", "Rediseño marca",
     "Antes/después: rediseño de marca que duplicó conversión",
     "Carrusel 6 slides. Slide 1: marca anterior. Slide 2-3: proceso y decisiones. Slide 4-5: marca nueva + aplicaciones. Slide 6: métrica (+180% conversión web). Estilo premium lima/negro.",
     "#Branding #Rediseño #Casos #IdentidadVisual",
     "¿Tu marca necesita evolución?",
     "Carrusel 1080×1350", "Programado"),
    (1, "Sáb", "2026-01-10", "Reel", "05 · Cultura", "Backstage sprint",
     "Inside ELEMENTAL: un día de trabajo con el equipo",
     "Reel 35s. Edición rápida tipo vlog. Muestra: reunión estrategia, design sprint, call con cliente, implementación IA. Audio instrumental. Texto overlay con momentos clave.",
     "#Backstage #Agencia #Cultura #Equipo",
     "Síguenos para más",
     "Reel 1080×1920", "Programado"),
    (1, "Dom", "2026-01-11", "Story + Caso", "04 · Casos", "Caso restaurante",
     "Caso de éxito: +180% reservas con WhatsApp IA",
     "Serie de 5 stories. 1: hook con resultado. 2: contexto del cliente. 3: problema. 4: solución implementada. 5: CTA con link a caso completo en bio. Stickers de encuesta y pregunta.",
     "—",
     "Toca el link en bio",
     "Stories 1080×1920", "Programado"),

    # ====== SEMANA 2 ======
    (2, "Lun", "2026-01-12", "Carrusel", "02 · Marketing", "Embudo ventas",
     "Embudo de ventas en 6 pasos (con métricas reales)",
     "Carrusel 8 slides. Diagrama embudo visual con números: 10k alcance → 500 clicks → 120 leads → 80 calificados → 30 propuestas → 12 cierres. Cada slide explica 1 etapa + qué medir. Slide final: CTA a diagnóstico.",
     "#Embudo #Ventas #Marketing #Funnel",
     "¿En qué etapa pierdes clientes?",
     "Carrusel 1080×1350", "Borrador"),
    (2, "Mar", "2026-01-13", "Reel", "01 · IA", "Generador cotizaciones IA",
     "Cómo generar cotizaciones en 3 segundos con IA",
     "Reel 30s. Demo pantalla: agente IA recibe brief, genera cotización completa en PDF, la envía por WhatsApp. Voz explicando flujo. Hook: 'Lo que tomaba 3 días, ahora 3 segundos.'",
     "#IA #Cotizaciones #Automatizacion #B2B",
     "Comenta COTIZA y te muestro",
     "Reel 1080×1920", "Borrador"),
    (2, "Mié", "2026-01-14", "Carrusel", "04 · Casos", "Mini auditoría inmobiliaria",
     "Mini auditoría: inmobiliaria que perdía 60% de leads",
     "Carrusel 7 slides. Slide 1: hook con número. Slides 2-5: 4 errores (sin CRM, WhatsApp compartido, sin seguimiento, sin calificación). Slide 6: solución con sistema. Slide 7: CTA diagnóstico.",
     "#Inmobiliaria #CRM #Leads #Casos",
     "¿Cuántos leads pierdes?",
     "Carrusel 1080×1350", "Borrador"),
    (2, "Jue", "2026-01-15", "Reel", "02 · Marketing", "Error SEO local",
     "El #1 error de SEO que mató tu tráfico local",
     "Reel 25s. Pantalla dividida: web sin optimizar vs web optimizada. Voz explica Google Business Profile + reseñas + structured data. Hook: 'Tu competencia ya lo hizo.'",
     "#SEO #Local #MarketingDigital #Google",
     "Comenta SEO y te envío checklist",
     "Reel 1080×1920", "Borrador"),
    (2, "Vie", "2026-01-16", "Carrusel", "03 · Antes/Después", "Web antes/después",
     "Antes/después: web que pasó de 2s a 0.8s de carga",
     "Carrusel 6 slides. Slide 1: web antigua con métricas. Slides 2-3: diagnóstico técnico. Slides 4-5: web nueva + optimizaciones. Slide 6: impacto en conversión. Mockups premium.",
     "#Web #Performance #Desarrollo #UX",
     "¿Tu web carga rápido?",
     "Carrusel 1080×1350", "Borrador"),
    (2, "Sáb", "2026-01-17", "Reel", "05 · Cultura", "Herramientas ELEMENTAL",
     "Las 7 herramientas que usamos en ELEMENTAL",
     "Reel 35s. Listado rápido con pantallas: Notion, Figma, Make, OpenAI, Vercel, Linear, Slack. Una por segundo con texto overlay. Hook: 'Stack 2026 para agencias tech.'",
     "#Herramientas #Stack #Productividad #Agencia",
     "¿Cuál usas tú?",
     "Reel 1080×1920", "Borrador"),
    (2, "Dom", "2026-01-18", "Story + Caso", "04 · Casos", "Caso inmobiliaria",
     "Caso: 320 leads/mes con sistema automatizado",
     "Serie de 5 stories. 1: hook con métrica. 2: contexto constructora. 3: problema embudo roto. 4: solución landings+CRM+IA. 5: CTA caso completo. Encuesta: '¿Captas leads así?'",
     "—",
     "Link en bio",
     "Stories 1080×1920", "Borrador"),

    # ====== SEMANA 3 ======
    (3, "Lun", "2026-01-19", "Carrusel", "01 · IA", "Tendencias IA 2026",
     "7 tendencias de IA para empresas en 2026",
     "Carrusel 8 slides. Una tendencia por slide: agentes autónomos, voice AI, RAG empresarial, multimodal, copilotos internos, IA generativa video, automatización no-code. Cada slide: tendencia + aplicación práctica + caso breve.",
     "#IA #Tendencias #2026 #Innovacion",
     "¿Cuál aplicarás primero?",
     "Carrusel 1080×1350", "Borrador"),
    (3, "Mar", "2026-01-20", "Reel", "01 · IA", "Voice AI demo",
     "Demo: agente de voz que atiende clientes 24/7",
     "Reel 30s. Llamada real simulada: cliente pregunta por horario y precios, agente IA responde natural, agenda cita. Pantalla dividida con波形. Hook: 'Esto no es futuro, es hoy.'",
     "#VoiceAI #IA #AtencionAutomatica",
     "Comenta VOZ y te envío demo",
     "Reel 1080×1920", "Borrador"),
    (3, "Mié", "2026-01-21", "Carrusel", "04 · Casos", "Mini auditoría servicios B2B",
     "Mini auditoría: empresa B2B que tardaba 3 días en cotizar",
     "Carrusel 7 slides. Slide 1: hook. Slides 2-5: diagnóstico (sin plantillas, 3 áreas, sin seguimiento, sin visibilidad). Slide 6: sistema con IA. Slide 7: resultado -85% tiempo + CTA.",
     "#B2B #Cotizaciones #Automatizacion #Casos",
     "¿Cuánto tardas en cotizar?",
     "Carrusel 1080×1350", "Borrador"),
    (3, "Jue", "2026-01-22", "Reel", "02 · Marketing", "Email marketing 2026",
     "Email marketing en 2026: lo que ya NO funciona",
     "Reel 25s. 3 cosas que ya no funcionan: newsletters genéricos, asuntos clickbait, frecuencia fija. 3 que sí: segmentación por comportamiento, IA en copy, automatización por evento.",
     "#EmailMarketing #Marketing #Digital",
     "Comenta EMAIL y te envío plantilla",
     "Reel 1080×1920", "Borrador"),
    (3, "Vie", "2026-01-23", "Carrusel", "03 · Antes/Después", "Dashboard antes/después",
     "Antes/después: de Excel caótico a dashboard en vivo",
     "Carrusel 6 slides. Slide 1: screenshot Excel desordenado. Slides 2-3: problema de visibilidad. Slides 4-5: dashboard nuevo con KPIs lima. Slide 6: impacto decisiones + tiempo ahorrado.",
     "#Dashboard #Datos #Analitica #Casos",
     "¿Tomas decisiones con datos?",
     "Carrusel 1080×1350", "Borrador"),
    (3, "Sáb", "2026-01-24", "Reel", "05 · Cultura", "Metodología ELEMENTAL",
     "Nuestra metodología en 5 fases (sin humo)",
     "Reel 35s. Explicación visual: 01 Estrategia → 02 Identidad → 03 Web → 04 Portafolio → 05 Redes & ventas. Cada fase con ícono y entregable claro. Hook: 'Así pasamos de idea a sistema.'",
     "#Metodologia #Agencia #Proceso",
     "Síguenos para más",
     "Reel 1080×1920", "Borrador"),
    (3, "Dom", "2026-01-25", "Story + Caso", "04 · Casos", "Caso servicios B2B",
     "Caso: -85% tiempo en cotizaciones con IA",
     "Serie de 5 stories. 1: métrica impactante. 2: contexto empresa. 3: problema cotizaciones lentas. 4: solución IA + CRM + dashboard. 5: CTA caso completo. Pregunta: '¿Cuánto tardas en cotizar?'",
     "—",
     "Link en bio",
     "Stories 1080×1920", "Borrador"),

    # ====== SEMANA 4 ======
    (4, "Lun", "2026-01-26", "Carrusel", "02 · Marketing", "SEO en 2026",
     "SEO en 2026: 9 cosas que cambiarán este año",
     "Carrusel 10 slides. Slide 1: hook con 9 puntos. Slides 2-9: una tendencia por slide (SGE, voice search, E-E-A-T, video SEO, AI overviews, local, core web vitals, contenido helpful, schema). Slide 10: CTA a auditoría SEO gratuita.",
     "#SEO #2026 #Marketing #Tendencias",
     "Comenta SEO y te audito",
     "Carrusel 1080×1350", "Borrador"),
    (4, "Mar", "2026-01-27", "Reel", "01 · IA", "Make vs n8n",
     "Make vs n8n: ¿cuál elegir para automatizar?",
     "Reel 30s. Comparación rápida: Make = visual + fácil, n8n = open source + flexible. Casos de uso para cada uno. Hook: 'Te ahorro 3 horas de investigación.'",
     "#Automatizacion #Make #n8n #NoCode",
     "¿Cuál usas tú?",
     "Reel 1080×1920", "Borrador"),
    (4, "Mié", "2026-01-28", "Carrusel", "04 · Casos", "Caso integral 360",
     "Caso 360: marca + web + IA + CRM en 8 semanas",
     "Carrusel 9 slides. Slide 1: hook. Slides 2-3: cliente y reto. Slides 4-7: 4 entregables (marca, web, IA, CRM) con screenshots. Slide 8: métricas a 90 días. Slide 9: testimonio + CTA.",
     "#Caso360 #Branding #Web #IA #CRM",
     "¿Tu empresa necesita esto?",
     "Carrusel 1080×1350", "Borrador"),
    (4, "Jue", "2026-01-29", "Reel", "02 · Marketing", "Hook para Reels",
     "3 hooks que disparan tus vistas en Reels",
     "Reel 25s. Ejemplos de hooks que funcionan: número impactante, pregunta incómoda, contradicción. Voz + texto overlay. Hook del propio reel: 'Probé 100 hooks y estos 3 funcionan.'",
     "#Reels #Hooks #Instagram #Marketing",
     "Comenta HOOK y te envío lista",
     "Reel 1080×1920", "Borrador"),
    (4, "Vie", "2026-01-30", "Carrusel", "03 · Antes/Después", "CRM antes/después",
     "Antes/después: de WhatsApp caótico a CRM en 2 semanas",
     "Carrusel 6 slides. Slide 1: WhatsApp desordenado. Slides 2-3: problema visibilidad. Slides 4-5: CRM organizado con pipeline. Slide 6: +60% conversión. Mockups lima/negro.",
     "#CRM #Ventas #Organizacion #Casos",
     "¿Usas CRM o WhatsApp?",
     "Carrusel 1080×1350", "Borrador"),
    (4, "Sáb", "2026-01-31", "Reel", "05 · Cultura", "Equipo remoto ELEMENTAL",
     "Cómo trabajamos remoto en ELEMENTAL (sin reuniones inútiles)",
     "Reel 35s. Día típico: async en Notion, daily 15min, deep work, ship. Tips de productividad. Hook: 'Trabajamos 6h y producimos más que 12h.'",
     "#Remoto #Productividad #Equipo #Async",
     "Síguenos para más",
     "Reel 1080×1920", "Borrador"),
    (4, "Dom", "2026-02-01", "Story + Caso", "04 · Casos", "Resumen mes",
     "Resumen del mes: 3 casos, 3 métricas, 3 aprendizajes",
     "Serie de 7 stories. 1: hook. 2-4: un caso por story con métrica. 5-6: aprendizajes clave. 7: CTA diagnóstico gratuito. Encuesta: '¿Cuál caso se parece a tu reto?'",
     "—",
     "Agenda diagnóstico",
     "Stories 1080×1920", "Borrador"),
]

# Write posts
for i, post in enumerate(posts, start=4):
    for j, val in enumerate(post, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        cell.fill = fill_surface
        cell.border = border_all
        cell.alignment = align_top_left
        # Style by column
        if j == 1:  # Semana
            cell.font = font_lima_bold
            cell.alignment = align_center
        elif j == 2:  # Día
            cell.font = Font(name="JetBrains Mono", size=10, color=LIMA)
            cell.alignment = align_center
        elif j == 3:  # Fecha
            cell.font = Font(name="JetBrains Mono", size=9, color=GRIS_CLARO)
            cell.alignment = align_center
        elif j == 4:  # Tipo
            cell.font = Font(name="Space Grotesk", size=10, bold=True, color=BLANCO)
            cell.alignment = align_center
            if "Carrusel" in val:
                cell.fill = PatternFill("solid", fgColor="2A2A0A")
            elif "Reel" in val:
                cell.fill = PatternFill("solid", fgColor="1A2A1A")
            elif "Story" in val:
                cell.fill = PatternFill("solid", fgColor="2A1A2A")
        elif j == 5:  # Pilar
            cell.font = font_mono
            cell.alignment = align_center
        elif j == 7:  # Título
            cell.font = Font(name="Space Grotesk", size=11, bold=True, color=BLANCO)
        elif j == 11:  # Formato
            cell.font = font_muted
        elif j == 12:  # Estado
            cell.font = Font(name="Inter", size=10, bold=True, color=LIMA if val == "Programado" else "FFA500")
            cell.alignment = align_center
        else:
            cell.font = font_body
    # Row height
    ws2.row_dimensions[i].height = 60

set_col_widths(ws2, [9, 7, 11, 14, 13, 18, 28, 50, 28, 22, 16, 12])

# Enable filter
ws2.auto_filter.ref = f"A3:L{3 + len(posts)}"

# ============= SHEET 3: FORMATOS =============
ws3 = wb.create_sheet("Formatos")
ws3.sheet_view.showGridLines = False
ws3.sheet_view.zoomScale = 110

# Dark bg
for r in range(1, 40):
    for c in range(1, 10):
        ws3.cell(row=r, column=c).fill = fill_negro

ws3["B2"] = "FORMATOS Y ESPECIFICACIONES"
ws3["B2"].font = Font(name="Space Grotesk", size=14, bold=True, color=BLANCO)
ws3["B2"].fill = fill_negro
ws3.merge_cells("B2:I2")

ws3["B3"] = "Especificaciones técnicas por formato. Todas las piezas respetan paleta ELEMENTAL."
ws3["B3"].font = Font(name="Inter", size=10, color="8A8A8A")
ws3["B3"].fill = fill_negro
ws3.merge_cells("B3:I3")

# Headers
fmt_headers = ["Formato", "Dimensiones", "Duración", "Pilar sugerido", "Estructura", "Tipografía", "Paleta", "CTA típico"]
for i, h in enumerate(fmt_headers, 2):
    cell = ws3.cell(row=5, column=i, value=h)
    cell.fill = fill_surface_2
    cell.font = font_header
    cell.alignment = align_center
    cell.border = border_all
ws3.row_dimensions[5].height = 28

# Format rows
formats = [
    ("Carrusel", "1080×1350 px", "6-10 slides", "01, 02, 03, 04",
     "Slide 1 hook con número o pregunta. Slides 2-N: 1 idea por slide con título corto + cuerpo breve. Slide final: CTA claro.",
     "Space Grotesk 700 para títulos. Inter 400 para cuerpo. JetBrains Mono para tags.",
     "Fondo #0D0D0D. Acentos #C8D400 (máx 10%). Texto #FFFFFF.",
     "Guarda / Comenta / Link en bio"),
    ("Reel", "1080×1920 px", "25-35 seg", "01, 02, 05",
     "Hook en primeros 3s (número o contradicción). Desarrollo 15-20s con 1 idea clara. CTA en últimos 5s. Texto overlay grande. Subtítulos siempre.",
     "Space Grotesk 700 overlay. Voz clara, audio trending si aplica.",
     "Fondo oscuro. Acento lima en texto overlay clave. Bordes limpios.",
     "Comenta [palabra clave]"),
    ("Story", "1080×1920 px", "5-7 stories", "04 (casos)",
     "Story 1: hook con métrica. Story 2-3: contexto y problema. Story 4-5: solución. Story final: CTA con sticker o link. Encuestas y preguntas para engagement.",
     "Space Grotesk 700 para hook. Inter 400 para cuerpo.",
     "Fondo #0D0D0D o foto con overlay negro 70%. Acento lima en métricas.",
     "Link en bio / Toca aquí"),
    ("Live / Webinar", "1920×1080 px", "30-45 min", "01, 02",
     "Estructura: 5min intro, 20min contenido educativo, 10min Q&A, 5min CTA. Mínimo 3 ejemplos prácticos. Material descargable al final.",
     "Slide deck con Space Grotesk 700 + Inter 400.",
     "Fondo oscuro, slides con mucho espacio negativo. Acento lima en datos clave.",
     "Agenda diagnóstico"),
    ("Post premium", "1080×1080 px", "1 imagen", "02, 05",
     "Frase estratégica o quote. Tipografía grande, mucho espacio negativo, máximo 15 palabras. Logo en esquina inferior.",
     "Space Grotesk 700 grande para frase. Logo pequeño en esquina.",
     "Fondo #0D0D0D. Texto #FFFFFF. Acento lima en máximo 1 palabra.",
     "Comparte si te resonó"),
]

for i, row in enumerate(formats, start=6):
    for j, val in enumerate(row, 2):
        cell = ws3.cell(row=i, column=j, value=val)
        cell.fill = fill_surface
        cell.border = border_all
        cell.alignment = align_top_left
        if j == 2:  # Formato name
            cell.font = Font(name="Space Grotesk", size=12, bold=True, color=LIMA)
            cell.alignment = align_center
        elif j == 3:  # Dimensions
            cell.font = Font(name="JetBrains Mono", size=10, color=BLANCO)
            cell.alignment = align_center
        else:
            cell.font = font_body
    ws3.row_dimensions[i].height = 90

set_col_widths(ws3, [3, 14, 16, 14, 16, 38, 24, 26, 22])

# ============= SHEET 4: PILARES Y TEMAS =============
ws4 = wb.create_sheet("Pilares y Temas")
ws4.sheet_view.showGridLines = False
ws4.sheet_view.zoomScale = 110

for r in range(1, 50):
    for c in range(1, 10):
        ws4.cell(row=r, column=c).fill = fill_negro

ws4["B2"] = "PILARES Y BANCO DE TEMAS"
ws4["B2"].font = Font(name="Space Grotesk", size=14, bold=True, color=BLANCO)
ws4["B2"].fill = fill_negro
ws4.merge_cells("B2:H2")

ws4["B3"] = "Banco de ideas por pilar para alimentar 12+ semanas de contenido sin bloqueo creativo."
ws4["B3"].font = Font(name="Inter", size=10, color="8A8A8A")
ws4["B3"].fill = fill_negro
ws4.merge_cells("B3:H3")

pillar_headers = ["Pilar", "Descripción", "Tema 1", "Tema 2", "Tema 3", "Tema 4", "Tema 5"]
for i, h in enumerate(pillar_headers, 2):
    cell = ws4.cell(row=5, column=i, value=h)
    cell.fill = fill_surface_2
    cell.font = font_header
    cell.alignment = align_center
    cell.border = border_all
ws4.row_dimensions[5].height = 28

pillars_data = [
    ("01 · IA para empresas",
     "Prompts, bots, automatizaciones, casos prácticos y tendencias aplicadas.",
     "5 prompts IA para WhatsApp",
     "Cómo montar voice AI en 1 día",
     "Agentes autónomos: guía 2026",
     "RAG empresarial: qué es y cómo usarlo",
     "IA generativa para contenido"),
    ("02 · Marketing y ventas",
     "Embudos, anuncios, contenido, SEO, redes y estrategias de conversión.",
     "Embudo de ventas en 6 pasos",
     "Meta Ads: errores comunes 2026",
     "SEO local: checklist completo",
     "Email marketing que convierte",
     "Hooks para Reels que funcionan"),
    ("03 · Antes / después",
     "Transformaciones de marca, web, piezas, campañas y procesos.",
     "Rediseño de marca +180% conversión",
     "Web 2s → 0.8s de carga",
     "Excel caótico → dashboard en vivo",
     "WhatsApp caótico → CRM organizado",
     "Cotización manual → IA automática"),
    ("04 · Casos reales",
     "Resultados, aprendizajes y mini auditorías de empresas.",
     "Mini auditoría restaurante",
     "Mini auditoría inmobiliaria",
     "Mini auditoría servicios B2B",
     "Caso 360 en 8 semanas",
     "Resumen mensual de resultados"),
    ("05 · Cultura ELEMENTAL",
     "Proceso creativo, metodología, equipo, herramientas y backstage.",
     "Stack de herramientas 2026",
     "Metodología en 5 fases",
     "Día de trabajo remoto",
     "Cómo trabajamos async",
     "Backstage de un sprint"),
]

for i, row in enumerate(pillars_data, start=6):
    for j, val in enumerate(row, 2):
        cell = ws4.cell(row=i, column=j, value=val)
        cell.fill = fill_surface
        cell.border = border_all
        cell.alignment = align_top_left
        if j == 2:  # Pilar name
            cell.font = Font(name="Space Grotesk", size=11, bold=True, color=LIMA)
        elif j == 3:  # Description
            cell.font = Font(name="Inter", size=10, italic=True, color=GRIS_CLARO)
        else:
            cell.font = font_body
    ws4.row_dimensions[i].height = 75

set_col_widths(ws4, [3, 22, 30, 22, 22, 22, 22, 22])

# ============= SHEET 5: MÉTRICAS =============
ws5 = wb.create_sheet("Métricas")
ws5.sheet_view.showGridLines = False
ws5.sheet_view.zoomScale = 110

for r in range(1, 40):
    for c in range(1, 12):
        ws5.cell(row=r, column=c).fill = fill_negro

ws5["B2"] = "MÉTRICAS Y SEGUIMIENTO"
ws5["B2"].font = Font(name="Space Grotesk", size=14, bold=True, color=BLANCO)
ws5["B2"].fill = fill_negro
ws5.merge_cells("B2:K2")

ws5["B3"] = "Tabla para registrar métricas semanales por publicación. Las fórmulas calculan totales y promedios automáticamente."
ws5["B3"].font = Font(name="Inter", size=10, color="8A8A8A")
ws5["B3"].fill = fill_negro
ws5["B3"].alignment = align_top_left
ws5.merge_cells("B3:K3")
ws5.row_dimensions[3].height = 30

# Headers
metric_headers = ["Publicación", "Tipo", "Alcance", "Impresiones", "Interacciones", "Guardados", "Comentarios", "Clicks", "Leads", "Tasa Engage %", "ROI"]
for i, h in enumerate(metric_headers, 2):
    cell = ws5.cell(row=5, column=i, value=h)
    cell.fill = fill_surface_2
    cell.font = font_header
    cell.alignment = align_center
    cell.border = border_all
ws5.row_dimensions[5].height = 32

# Sample rows (first 12 posts)
sample_posts = [
    ("W1 · Carrusel prompts IA", "Carrusel"),
    ("W1 · Reel bot WhatsApp", "Reel"),
    ("W1 · Carrusel auditoría restaurante", "Carrusel"),
    ("W1 · Reel error Meta Ads", "Reel"),
    ("W1 · Carrusel rediseño marca", "Carrusel"),
    ("W1 · Reel backstage", "Reel"),
    ("W1 · Story caso restaurante", "Story"),
    ("W2 · Carrusel embudo ventas", "Carrusel"),
    ("W2 · Reel cotizaciones IA", "Reel"),
    ("W2 · Carrusel auditoría inmobiliaria", "Carrusel"),
    ("W2 · Reel SEO local", "Reel"),
    ("W2 · Carrusel web antes/después", "Carrusel"),
]
for i, (name, tipo) in enumerate(sample_posts, start=6):
    ws5.cell(row=i, column=2, value=name).font = Font(name="Space Grotesk", size=10, color=BLANCO)
    ws5.cell(row=i, column=3, value=tipo).font = font_mono
    # Empty cells for user input
    for c in range(4, 11):
        ws5.cell(row=i, column=c, value=0).font = Font(name="JetBrains Mono", size=10, color=LIMA)
    # Formula: Tasa Engage % = Interacciones / Alcance * 100
    ws5.cell(row=i, column=11, value=f"=IFERROR(F{i}/D{i}*100, 0)").font = Font(name="JetBrains Mono", size=10, color=LIMA)
    # ROI formula placeholder (Leads * valor estimado / inversión)
    ws5.cell(row=i, column=12, value=f"=IFERROR(J{i}*50/100, 0)").font = Font(name="JetBrains Mono", size=10, color=LIMA)
    
    for c in range(2, 13):
        cell = ws5.cell(row=i, column=c)
        cell.fill = fill_surface
        cell.border = border_all
        if c == 2:
            cell.alignment = align_left
        elif c == 3:
            cell.alignment = align_center
        else:
            cell.alignment = align_center
            if c >= 4 and c <= 10:
                cell.number_format = "#,##0"
            elif c == 11:
                cell.number_format = "0.0\"%\""
            elif c == 12:
                cell.number_format = "$#,##0"

# Total row
total_row = 6 + len(sample_posts)
ws5.cell(row=total_row, column=2, value="TOTAL").font = Font(name="Space Grotesk", size=11, bold=True, color=BLANCO)
ws5.cell(row=total_row, column=2).fill = fill_surface_2
ws5.cell(row=total_row, column=2).alignment = align_center
ws5.cell(row=total_row, column=2).border = border_all

ws5.cell(row=total_row, column=3, value="—").font = font_muted
ws5.cell(row=total_row, column=3).fill = fill_surface_2
ws5.cell(row=total_row, column=3).alignment = align_center
ws5.cell(row=total_row, column=3).border = border_all

for c in range(4, 11):
    col_letter = get_column_letter(c)
    ws5.cell(row=total_row, column=c, value=f"=SUM({col_letter}6:{col_letter}{total_row-1})")
    cell = ws5.cell(row=total_row, column=c)
    cell.font = Font(name="JetBrains Mono", size=11, bold=True, color=LIMA)
    cell.fill = fill_surface_2
    cell.alignment = align_center
    cell.border = border_all
    cell.number_format = "#,##0"

# Average engagement
ws5.cell(row=total_row, column=11, value=f"=IFERROR(F{total_row}/D{total_row}*100, 0)")
cell = ws5.cell(row=total_row, column=11)
cell.font = Font(name="JetBrains Mono", size=11, bold=True, color=LIMA)
cell.fill = fill_surface_2
cell.alignment = align_center
cell.border = border_all
cell.number_format = "0.0\"%\""

# Total ROI
ws5.cell(row=total_row, column=12, value=f"=SUM(L6:L{total_row-1})")
cell = ws5.cell(row=total_row, column=12)
cell.font = Font(name="JetBrains Mono", size=11, bold=True, color=LIMA)
cell.fill = fill_surface_2
cell.alignment = align_center
cell.border = border_all
cell.number_format = "$#,##0"

ws5.row_dimensions[total_row].height = 28

# KPIs objetivo
kpi_row = total_row + 3
ws5.cell(row=kpi_row, column=2, value="KPIs OBJETIVO MENSUAL").font = font_subheader
ws5.cell(row=kpi_row, column=2).fill = fill_negro

kpis_table = [
    ("Alcance total", f"=D{total_row}", "≥ 80,000"),
    ("Impresiones totales", f"=E{total_row}", "≥ 150,000"),
    ("Guardados totales", f"=G{total_row}", "≥ 600"),
    ("Leads generados", f"=J{total_row}", "≥ 30"),
    ("Tasa de engagement", f"=K{total_row}", "≥ 3.5%"),
    ("ROI estimado (USD)", f"=L{total_row}", "≥ $1,500"),
]
for i, (k, formula, target) in enumerate(kpis_table, start=kpi_row+1):
    ws5.cell(row=i, column=2, value=k).font = font_body
    ws5.cell(row=i, column=2).fill = fill_surface
    ws5.cell(row=i, column=2).border = border_all
    ws5.cell(row=i, column=2).alignment = align_left
    ws5.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)
    
    ws5.cell(row=i, column=6, value=formula).font = font_lima_bold
    ws5.cell(row=i, column=6).fill = fill_surface
    ws5.cell(row=i, column=6).border = border_all
    ws5.cell(row=i, column=6).alignment = align_center
    
    ws5.cell(row=i, column=7, value=target).font = font_muted
    ws5.cell(row=i, column=7).fill = fill_surface
    ws5.cell(row=i, column=7).border = border_all
    ws5.cell(row=i, column=7).alignment = align_center

set_col_widths(ws5, [3, 30, 14, 12, 14, 14, 14, 14, 12, 12, 14, 14])

# ============= SAVE =============
output_path = "/home/z/my-project/download/elemental/ELEMENTAL_Calendario_Contenido.xlsx"
wb.save(output_path)
print(f"XLSX saved: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Posts in calendar: {len(posts)}")
